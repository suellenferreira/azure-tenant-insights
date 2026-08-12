"""
Excel inventory writer using openpyxl.

Generates a formatted, multi-sheet Excel workbook from collected scan data.

Sheets produced:
  Overview         - KPI summary + top resource types + Advisor by pillar
    Subscriptions    - Resource inventory aggregation by subscription/RG/location/type
    AllResources     - Flat table of ALL resources with common ARG/ARM columns
    [ResourceType]   - One sheet per resource type with declarative enrichment
  AdvisorFindings  - All Advisor recommendations with WAF pillar
  PolicyCompliance - Non-compliant resources
  ResourceHealth   - Degraded/unavailable resources
  DeprecatedRes    - Resources matching retirement announcements
  MisconfigFindgs  - Known misconfiguration findings
  SecurityAssessmt - Defender for Cloud assessments (optional)
  DefenderPosture  - Defender plan enablement per subscription (optional)
  DefenderServersCoverage - Per-resource VM/VMSS/Arc Defender coverage (optional)
  DefenderCoverageGap - Unprotected billable units & cost to protect (optional)
  Costs            - Cost by RG/Service (optional)
"""

import json
import logging
import os
from typing import Any, Dict, List

from writers.safety import excel_safe_data

logger = logging.getLogger(__name__)

# Excel column color constants
COLOR_HEADER_BG = "1F4E79"
COLOR_HEADER_FG = "FFFFFF"
COLOR_CRITICAL = "C00000"
COLOR_HIGH = "FF4444"
COLOR_MEDIUM = "FFC000"
COLOR_LOW = "FFFF99"
COLOR_OK = "70AD47"
COLOR_SECTION_BG = "D6E4F0"
COLOR_ALT_ROW = "F2F7FB"


def _env_int(name: str, default: int) -> int:
    try:
        return int(os.getenv(name, str(default)))
    except (TypeError, ValueError):
        return default


# ── Worksheet-count governance ───────────────────────────────────────────────
# Excel supports at most 255 worksheets — this is a hard cap. Resource types beyond
# the available budget are still represented in the AllResources tab (no own sheet).
EXCEL_MAX_SHEETS = 255

# Scope-aware warning thresholds: emit a warning once the number of resource-type
# sheets reaches these values (navigation becomes cumbersome). Env-overridable.
SCOPE_WARN_THRESHOLDS = {
    "subscription": _env_int("ATI_SHEET_WARN_SUBSCRIPTION", 40),
    "management-group": _env_int("ATI_SHEET_WARN_MANAGEMENT_GROUP", 60),
    "tenant": _env_int("ATI_SHEET_WARN_TENANT", 75),
}
# Hard warning once the sheet count approaches the Excel ceiling.
SHEET_WARN_HARD = _env_int("ATI_SHEET_WARN_HARD", 200)

# Fixed sheet names — reserved so type sheets never collide with them.
RESERVED_SHEET_NAMES = {
    "Overview", "Index", "Classification", "ModernizationSignals", "ResiliencyEvidence", "Subscriptions",
    "AllResources", "AdvisorFindings", "PolicyCompliance", "ResourceHealth",
    "DeprecatedResources", "MisconfigFindings", "SecurityAssessments",
    "DefenderCostEstimate", "DefenderPosture", "DefenderServersCoverage",
    "DefenderCoverageGap", "Costs",
}

CORE_RESOURCE_TYPES = [
    "microsoft.compute/virtualmachines",
    "microsoft.compute/disks",
    "microsoft.network/networkinterfaces",
    "microsoft.network/publicipaddresses",
    "microsoft.network/networksecuritygroups",
    "microsoft.network/virtualnetworks",
    "microsoft.storage/storageaccounts",
    "microsoft.web/sites",
    "microsoft.web/serverfarms",
    "microsoft.keyvault/vaults",
    "microsoft.sql/servers",
    "microsoft.sql/servers/databases",
    "microsoft.network/privateendpoints",
]


def write_excel(scan_data: dict, output_path: str) -> None:
    """Main entry point — builds the full Excel workbook."""
    import openpyxl

    # Resource names, tags, and collector fields may be attacker-controlled.
    # Treat formula-looking strings as literal workbook values.
    scan_data = excel_safe_data(scan_data)
    logger.info(f"Building Excel workbook: {output_path}")
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # Remove the default empty sheet

    options = scan_data.get("options", {})
    include_tags = options.get("include_tags", False)

    # Single source-of-truth resource_type → sheet-name map (namespace-aware,
    # collision-free), shared across the type sheets and the AllResources tab.
    sheet_map = _prepare_type_sheet_map(scan_data)

    _write_overview_sheet(wb, scan_data)
    _write_classification_sheet(wb, scan_data, sheet_map)
    _write_modernization_sheet(wb, scan_data)
    _write_resiliency_evidence_sheet(wb, scan_data)
    _write_subscriptions_sheet(wb, scan_data)
    _write_all_resources_sheet(wb, scan_data, include_tags, sheet_map)
    _write_resource_type_sheets(wb, scan_data, include_tags, sheet_map)
    _write_advisor_sheet(wb, scan_data)
    _write_policy_sheet(wb, scan_data)
    _write_health_sheet(wb, scan_data)
    _write_deprecated_sheet(wb, scan_data)
    _write_misconfig_sheet(wb, scan_data)

    if scan_data.get("defender_data"):
        _write_defender_sheet(wb, scan_data)
        _write_defender_cost_estimate_sheet(wb, scan_data)

    if scan_data.get("defender_posture"):
        _write_defender_posture_sheet(wb, scan_data)
        _write_defender_servers_coverage_sheet(wb, scan_data)
        _write_defender_coverage_gap_sheet(wb, scan_data)

    if scan_data.get("costs_data"):
        _write_costs_sheet(wb, scan_data)

    # Navigation: Index sheet (positioned right after Overview) + per-sheet
    # "back to Index" links.
    _write_index_sheet(wb)
    _add_back_links(wb)

    wb.save(output_path)
    logger.info(f"Excel workbook saved: {output_path}")


def _write_resiliency_evidence_sheet(wb, scan_data: dict):
    """Write observed regional resiliency evidence with classification context."""
    from openpyxl.styles import Alignment, Font, PatternFill
    from processors.resiliency import build_resiliency_assessment

    assessment = build_resiliency_assessment(scan_data)
    if not assessment.get("available"):
        return

    environment = assessment.get("environment", {})
    ws = wb.create_sheet("ResiliencyEvidence")
    ws.sheet_view.showGridLines = False
    ws.merge_cells("A1:F1")
    ws["A1"] = "Resiliency Evidence — Current Environment + Control Signals"
    ws["A1"].font = Font(bold=True, size=15, color="FFFFFF")
    ws["A1"].fill = PatternFill("solid", fgColor=COLOR_HEADER_BG)
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 28
    ws.merge_cells("A2:F2")
    ws["A2"] = "Observed regional distribution and zone signals; not a workload-level resiliency certification."
    ws["A2"].font = Font(size=9, italic=True, color="777777")

    ws.merge_cells("A4:F4")
    ws["A4"] = "REGIONAL EVIDENCE BY CLASSIFICATION"
    ws["A4"].font = Font(bold=True, size=11, color="FFFFFF")
    ws["A4"].fill = PatternFill("solid", fgColor="1F4E79")
    ws["A4"].alignment = Alignment(horizontal="center")
    headers = ["Region", "Resources", "%", "Subscriptions", "Business Pillars", "Technical Categories"]
    for column, header in enumerate(headers, 1):
        ws.cell(5, column).value = header
    _header_style(ws, 5, len(headers))
    _auto_filter(ws, 5, len(headers))
    _freeze(ws, 6)
    for row_number, region in enumerate(environment.get("region_distribution", []), 6):
        ws.cell(row_number, 1).value = region.get("region", "")
        ws.cell(row_number, 2).value = region.get("resources", 0)
        ws.cell(row_number, 3).value = region.get("percentage", 0)
        ws.cell(row_number, 4).value = region.get("subscriptions", 0)
        ws.cell(row_number, 5).value = ", ".join((region.get("business_pillars") or {}).keys())
        ws.cell(row_number, 6).value = ", ".join((region.get("technical_categories") or {}).keys())

    summary_row = 7 + len(environment.get("region_distribution", []))
    ws.merge_cells(start_row=summary_row, start_column=1, end_row=summary_row, end_column=2)
    ws.cell(summary_row, 1).value = "TECHNICAL CATEGORY DISTRIBUTION"
    ws.cell(summary_row, 1).font = Font(bold=True, size=11, color="FFFFFF")
    ws.cell(summary_row, 1).fill = PatternFill("solid", fgColor="1F4E79")
    ws.cell(summary_row + 1, 1).value = "Technical Category"
    ws.cell(summary_row + 1, 2).value = "Resources"
    _header_style(ws, summary_row + 1, 2)
    for row_number, (category, count) in enumerate(
        sorted((environment.get("technical_category_distribution") or {}).items(), key=lambda item: -item[1]),
        summary_row + 2,
    ):
        ws.cell(row_number, 1).value = category
        ws.cell(row_number, 2).value = count

    def _write_region_matrix(start_row: int, title: str, label_header: str, key: str) -> int:
        regions = environment.get("region_distribution", []) or []
        region_names = [region.get("region", "") for region in regions]
        labels = sorted({
            label
            for region in regions
            for label in (region.get(key) or {}).keys()
        })
        last_col = max(2, len(region_names) + 1)
        ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=last_col)
        ws.cell(start_row, 1).value = title
        ws.cell(start_row, 1).font = Font(bold=True, size=11, color="FFFFFF")
        ws.cell(start_row, 1).fill = PatternFill("solid", fgColor="1F4E79")
        ws.cell(start_row, 1).alignment = Alignment(horizontal="center")

        header_row = start_row + 1
        ws.cell(header_row, 1).value = label_header
        for column, region_name in enumerate(region_names, 2):
            ws.cell(header_row, column).value = region_name
        _header_style(ws, header_row, last_col)

        for offset, label in enumerate(labels, 1):
            row_number = header_row + offset
            ws.cell(row_number, 1).value = label
            for column, region in enumerate(regions, 2):
                ws.cell(row_number, column).value = (region.get(key) or {}).get(label, 0)
        return header_row + len(labels) + 2

    matrix_start = summary_row + 3 + len(environment.get("technical_category_distribution") or {})
    matrix_start = _write_region_matrix(
        matrix_start, "RESILIENCY SNAPSHOT BY SERVICE MODEL", "Service Model", "service_models"
    )
    matrix_start = _write_region_matrix(
        matrix_start, "RESILIENCY SNAPSHOT BY BUSINESS PILLAR", "Business Pillar", "business_pillars"
    )
    _write_region_matrix(
        matrix_start, "RESILIENCY SNAPSHOT BY TECHNICAL CATEGORY", "Technical Category", "technical_categories"
    )
    _col_widths(ws, [34, 14, 14, 16, 42, 55, 18, 18, 18, 18, 18, 18])


# ─────────────────────────────────────────────────────────────────────────────
# Shared helpers
# ─────────────────────────────────────────────────────────────────────────────

def _header_style(ws, row_num: int, num_cols: int):
    from openpyxl.styles import Alignment, Font, PatternFill
    fill = PatternFill("solid", fgColor=COLOR_HEADER_BG)
    font = Font(bold=True, color=COLOR_HEADER_FG, size=11)
    align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=row_num, column=col)
        cell.fill = fill
        cell.font = font
        cell.alignment = align


def _auto_filter(ws, header_row: int, num_cols: int):
    from openpyxl.utils import get_column_letter
    ws.auto_filter.ref = f"A{header_row}:{get_column_letter(num_cols)}{header_row}"


def _freeze(ws, freeze_row: int = 2):
    ws.freeze_panes = f"A{freeze_row}"


def _col_widths(ws, widths: List[int]):
    from openpyxl.utils import get_column_letter
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def _severity_fill(severity: str):
    from openpyxl.styles import PatternFill
    colors = {
        "Critical": COLOR_CRITICAL,
        "High": COLOR_HIGH,
        "Medium": COLOR_MEDIUM,
        "Low": COLOR_LOW,
    }
    color = colors.get(severity)
    return PatternFill("solid", fgColor=color) if color else None


def _subscription_name_map(scan_data: dict) -> Dict[str, str]:
    return {
        sub.get("subscriptionId", ""): sub.get("displayName", "")
        for sub in scan_data.get("subscriptions", [])
    }


def _resource_display_name(resource_type: str, inventory_config: dict) -> str:
    from processors.normalizer import clean_resource_type

    type_config = inventory_config.get("resource_types", {}).get(resource_type, {})
    return type_config.get("display_name") or clean_resource_type(resource_type)


def _sheet_name_for_type(resource_type: str, inventory_config: dict) -> str:
    from processors.normalizer import excel_safe_sheet_name

    return excel_safe_sheet_name(_resource_display_name(resource_type, inventory_config))


# ── Namespace / category helpers ─────────────────────────────────────────────
_PROVIDER_ABBREV = {
    "microsoft.compute": "Cmp",
    "microsoft.network": "Net",
    "microsoft.storage": "Stor",
    "microsoft.web": "Web",
    "microsoft.keyvault": "Kv",
    "microsoft.sql": "Sql",
    "microsoft.connectedvmwarevsphere": "VMw",
    "microsoft.hybridcompute": "Arc",
    "microsoft.azurestackhci": "Hci",
    "microsoft.scvmm": "Scvmm",
    "microsoft.azuremigrate": "Mig",
    "microsoft.migrate": "Mig",
}

_HYBRID_PROVIDERS = {
    "microsoft.connectedvmwarevsphere",
    "microsoft.hybridcompute",
    "microsoft.azurestackhci",
    "microsoft.scvmm",
    "microsoft.hybridcontainerservice",
}
_MIGRATE_PROVIDERS = {
    "microsoft.azuremigrate",
    "microsoft.migrate",
    "microsoft.offazure",
    "microsoft.datareplication",
}


def _provider_of(resource_type: str) -> str:
    return resource_type.split("/", 1)[0].lower()


def _provider_token(resource_type: str) -> str:
    """Short, human-friendly abbreviation for a resource provider namespace."""
    provider = _provider_of(resource_type)
    if provider in _PROVIDER_ABBREV:
        return _PROVIDER_ABBREV[provider]
    tail = provider.split(".")[-1] if "." in provider else provider
    return (tail[:3] or "Res").title()


def _resource_category(resource_type: str) -> str:
    """Classify a resource type as Azure-native, hybrid/Arc, or Migrate."""
    provider = _provider_of(resource_type)
    if provider in _HYBRID_PROVIDERS:
        return "Hybrid / Arc"
    if provider in _MIGRATE_PROVIDERS:
        return "Migrate"
    return "Azure Native"


def _scan_scope(scan_data: dict) -> str:
    """Resolve the scan scope; fall back to inferring it from subscription count."""
    meta = scan_data.get("metadata", {})
    scope = meta.get("scan_scope")
    if scope in SCOPE_WARN_THRESHOLDS:
        return scope
    n_subs = meta.get("subscription_count") or len(scan_data.get("subscriptions", []))
    if n_subs <= 1:
        return "subscription"
    if n_subs <= 5:
        return "management-group"
    return "tenant"


def _build_sheet_name_map(sorted_types, inventory_config, budget, reserved_names) -> Dict[str, str]:
    """Single source of truth: resource_type → unique, collision-free sheet name.

    A namespace token is prepended ONLY when the bare display name would collide
    with another provider's type or with a reserved fixed-sheet name, keeping clean
    names for the common case while eliminating the same-suffix collision bug.
    """
    from processors.normalizer import excel_safe_sheet_name

    eligible = sorted_types[:budget]

    base_of: Dict[str, str] = {}
    base_counts: Dict[str, int] = {}
    for rtype, _rows in eligible:
        base = excel_safe_sheet_name(_resource_display_name(rtype, inventory_config))
        base_of[rtype] = base
        base_counts[base.lower()] = base_counts.get(base.lower(), 0) + 1

    used = {n.lower() for n in reserved_names}
    mapping: Dict[str, str] = {}
    for rtype, _rows in eligible:
        base = base_of[rtype]
        if base_counts[base.lower()] > 1 or base.lower() in used:
            name = excel_safe_sheet_name(f"{_provider_token(rtype)}-{base}")
        else:
            name = base
        final = name
        n = 2
        while final.lower() in used:
            suffix = f"~{n}"
            final = excel_safe_sheet_name(final[: 31 - len(suffix)] + suffix)
            n += 1
        used.add(final.lower())
        mapping[rtype] = final
    return mapping


def _reserved_sheet_count(scan_data: dict) -> int:
    """Number of fixed (non-type) sheets that will be created for this scan."""
    count = 11  # Overview, Index, Classification, ModernizationSignals,
    #             Subscriptions, AllResources, Advisor, Policy, Health,
    #             Deprecated, Misconfig
    if scan_data.get("defender_data"):
        count += 2  # SecurityAssessments + DefenderCostEstimate
    if scan_data.get("defender_posture"):
        count += 3  # DefenderPosture + DefenderServersCoverage + DefenderCoverageGap
    if scan_data.get("costs_data"):
        count += 1  # Costs
    return count


def _prepare_type_sheet_map(scan_data: dict) -> Dict[str, str]:
    """Build the shared type→sheet map and emit scope-aware capacity warnings."""
    from collectors.resources import load_enrichment_config

    resources_by_type = scan_data.get("resources_by_type", {})
    inventory_config = load_enrichment_config()
    sorted_types = _inventory_type_items(resources_by_type)

    reserved = _reserved_sheet_count(scan_data)
    budget = max(1, EXCEL_MAX_SHEETS - reserved)

    sheet_map = _build_sheet_name_map(
        sorted_types, inventory_config, budget, RESERVED_SHEET_NAMES
    )

    n_types = len(sorted_types)
    n_sheets = len(sheet_map)
    scope = _scan_scope(scan_data)
    threshold = SCOPE_WARN_THRESHOLDS.get(scope, 50)

    if n_sheets >= threshold:
        logger.warning(
            f"Workbook will contain {n_sheets} resource-type sheets "
            f"(scan scope='{scope}', warning threshold={threshold}). Navigation may be "
            f"cumbersome — use the Index sheet to jump between tabs."
        )
    if n_sheets >= SHEET_WARN_HARD:
        logger.warning(
            f"Resource-type sheet count ({n_sheets}) is approaching the Excel hard "
            f"limit of {EXCEL_MAX_SHEETS} sheets."
        )
    if n_types > budget:
        logger.warning(
            f"{n_types - budget} resource type(s) exceed the Excel {EXCEL_MAX_SHEETS}-sheet "
            f"limit and will appear only in the AllResources tab (no dedicated sheet)."
        )
    return sheet_map


def _inventory_type_items(resources_by_type: dict) -> List[tuple]:
    """Core resource types first, then remaining types by resource count."""
    selected = []
    seen = set()

    for resource_type in CORE_RESOURCE_TYPES:
        resources = resources_by_type.get(resource_type)
        if resources:
            selected.append((resource_type, resources))
            seen.add(resource_type)

    remaining = sorted(
        ((rtype, rows) for rtype, rows in resources_by_type.items() if rtype not in seen),
        key=lambda item: -len(item[1]),
    )
    return selected + remaining


def _enriched_columns(resource_type: str, resources: List[dict], inventory_config: dict) -> List[str]:
    cols = set()
    for resource in resources:
        cols.update(k for k in resource.keys() if k.startswith("enriched_"))

    type_config = inventory_config.get("resource_types", {}).get(resource_type, {})
    configured = [
        f"enriched_{field.get('column')}"
        for field in type_config.get("promoted_fields", [])
        if field.get("column")
    ]
    ordered = [col for col in configured if col in cols]
    extras = sorted(cols.difference(ordered))
    return ordered + extras


def _sku_value(resource: dict, key: str) -> str:
    from processors.normalizer import safe_str

    sku = resource.get("sku") or {}
    if isinstance(sku, dict):
        return safe_str(sku.get(key, ""))
    return safe_str(sku) if key == "name" else ""


def _common_resource_value(resource: dict, key: str) -> Any:
    properties = resource.get("properties") or {}
    identity = resource.get("identity") or {}
    if key == "provisioningState" and isinstance(properties, dict):
        return properties.get("provisioningState", "")
    if key == "createdTime" and isinstance(properties, dict):
        return (
            properties.get("timeCreated")
            or properties.get("createdTime")
            or properties.get("createdAt")
            or properties.get("dateCreated")
            or ""
        )
    if key == "identityType" and isinstance(identity, dict):
        return identity.get("type", "")
    return ""


# ─────────────────────────────────────────────────────────────────────────────
# Sheet writers
# ─────────────────────────────────────────────────────────────────────────────

def _write_overview_sheet(wb, scan_data: dict):
    from openpyxl.styles import Alignment, Font, PatternFill, Border, Side

    ws = wb.create_sheet("Overview", 0)
    ws.sheet_view.showGridLines = False
    ws.page_setup.paperSize = ws.PAPERSIZE_LETTER
    ws.page_margins.left = 0.5
    ws.page_margins.right = 0.5
    ws.page_margins.top = 0.75
    ws.page_margins.bottom = 0.75

    meta = scan_data.get("metadata", {})
    summary = scan_data.get("summary_metrics", {})

    # Dark theme header bar
    ws.merge_cells("A1:P1")
    title_cell = ws["A1"]
    title_cell.value = "Azure Tenant Insights — Environment Overview"
    title_cell.font = Font(bold=True, size=18, color="FFFFFF")
    title_cell.fill = PatternFill("solid", fgColor=COLOR_HEADER_BG)
    title_cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 42

    # Metadata section with subtle background
    meta_fill = PatternFill("solid", fgColor="E7E6E6")
    meta_font = Font(bold=True, size=10)
    ws["A2"] = "Scan Date"
    ws["A2"].fill = meta_fill
    ws["A2"].font = meta_font
    ws["B2"] = meta.get("scan_timestamp", "")[:10]
    ws["B2"].fill = PatternFill("solid", fgColor="F5F5F5")

    ws["C2"] = "Tenant ID"
    ws["C2"].fill = meta_fill
    ws["C2"].font = meta_font
    ws["D2"] = meta.get("tenant_id", "N/A")
    ws["D2"].fill = PatternFill("solid", fgColor="F5F5F5")

    ws["E2"] = "Cloud"
    ws["E2"].fill = meta_fill
    ws["E2"].font = meta_font
    ws["F2"] = meta.get("cloud", "AzurePublicCloud")
    ws["F2"].fill = PatternFill("solid", fgColor="F5F5F5")

    ws["A3"] = "Tenant Name"
    ws["A3"].fill = meta_fill
    ws["A3"].font = meta_font
    ws.merge_cells("B3:F3")
    ws["B3"] = meta.get("tenant_name", "N/A")
    ws["B3"].fill = PatternFill("solid", fgColor="F5F5F5")
    ws.row_dimensions[3].height = 18

    # Section 1: Key Metrics (with dark header)
    section_header_fill = PatternFill("solid", fgColor="1F4E79")
    section_header_font = Font(bold=True, size=11, color="FFFFFF")

    ws.merge_cells("A4:B4")
    ws["A4"].value = "KEY METRICS"
    ws["A4"].fill = section_header_fill
    ws["A4"].font = section_header_font
    ws["A4"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[4].height = 24

    kpi_rows = [
        ("Total Resources", summary.get("total_resources", 0), COLOR_SECTION_BG),
        ("Subscriptions", summary.get("total_subscriptions", 0), COLOR_SECTION_BG),
        ("Resource Types", summary.get("total_resource_types", 0), COLOR_SECTION_BG),
        ("Active Regions", summary.get("total_regions", 0), COLOR_SECTION_BG),
        ("Advisor Recommendations", summary.get("total_advisor_recommendations", 0), COLOR_MEDIUM),
        ("Policy Non-Compliant", summary.get("total_non_compliant_policies", 0), COLOR_MEDIUM),
        ("Deprecated Resources", summary.get("total_deprecated", 0), COLOR_HIGH),
        ("Critical Findings", summary.get("critical_findings_count", 0), COLOR_CRITICAL),
        ("Tag Coverage", f"{summary.get('tag_coverage_pct', 0)}%", COLOR_SECTION_BG),
        ("Health Issues", summary.get("total_health_issues", 0), COLOR_MEDIUM),
    ]

    for idx, (label, value, color) in enumerate(kpi_rows):
        row = 5 + idx
        label_cell = ws.cell(row=row, column=1)
        label_cell.value = label
        label_cell.font = Font(bold=True, size=10)
        label_cell.fill = PatternFill("solid", fgColor="F5F5F5")
        label_cell.alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[row].height = 22

        value_cell = ws.cell(row=row, column=2)
        value_cell.value = value
        value_cell.fill = PatternFill("solid", fgColor=color)
        value_cell.font = Font(bold=True, size=11)
        value_cell.alignment = Alignment(horizontal="center", vertical="center")
        value_cell.border = Border(
            left=Side(style="thin", color="D3D3D3"),
            right=Side(style="thin", color="D3D3D3"),
            top=Side(style="thin", color="D3D3D3"),
            bottom=Side(style="thin", color="D3D3D3"),
        )

    # Section 1b: Resource Origin (Azure-native vs hybrid/Arc vs Migrate)
    origin_counts = {"Azure Native": 0, "Hybrid / Arc": 0, "Migrate": 0}
    for _rt, _rows in scan_data.get("resources_by_type", {}).items():
        origin_counts[_resource_category(_rt)] += len(_rows)

    ws.merge_cells("A15:B15")
    ws["A15"].value = "RESOURCE ORIGIN"
    ws["A15"].fill = section_header_fill
    ws["A15"].font = section_header_font
    ws["A15"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[15].height = 22

    _origin_palette = {
        "Azure Native": COLOR_SECTION_BG,
        "Hybrid / Arc": COLOR_MEDIUM,
        "Migrate": "D9D2E9",
    }
    for _i, _label in enumerate(("Azure Native", "Hybrid / Arc", "Migrate")):
        _r = 16 + _i
        _lc = ws.cell(row=_r, column=1)
        _lc.value = _label
        _lc.font = Font(bold=True, size=10)
        _lc.fill = PatternFill("solid", fgColor="F5F5F5")
        _lc.alignment = Alignment(horizontal="left", vertical="center")
        _vc = ws.cell(row=_r, column=2)
        _vc.value = origin_counts[_label]
        _vc.fill = PatternFill("solid", fgColor=_origin_palette[_label])
        _vc.font = Font(bold=True, size=11)
        _vc.alignment = Alignment(horizontal="center", vertical="center")
        _vc.border = Border(
            left=Side(style="thin", color="D3D3D3"),
            right=Side(style="thin", color="D3D3D3"),
            top=Side(style="thin", color="D3D3D3"),
            bottom=Side(style="thin", color="D3D3D3"),
        )
        ws.row_dimensions[_r].height = 20

    # Section 1c: Service Model summary (assessment lens: IaaS/PaaS/SaaS/...)
    from processors.classifier import classify_resource_type, service_model_order

    _sm_counts: Dict[str, int] = {}
    for _rt, _rows in scan_data.get("resources_by_type", {}).items():
        _sm = classify_resource_type(_rt)["service_model"]
        _sm_counts[_sm] = _sm_counts.get(_sm, 0) + len(_rows)
    _sm_labels = [m for m in service_model_order() if _sm_counts.get(m)]
    _sm_labels += [m for m in _sm_counts if m not in _sm_labels]

    ws.merge_cells("D15:E15")
    ws["D15"].value = "SERVICE MODEL"
    ws["D15"].fill = section_header_fill
    ws["D15"].font = section_header_font
    ws["D15"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[15].height = 22

    _sm_palette = {
        "IaaS": COLOR_SECTION_BG,
        "PaaS": "D6E4F0",
        "SaaS": "E2EFDA",
        "Hybrid": COLOR_MEDIUM,
        "Supporting Services": "FCE4D6",
        "Other": "F2F2F2",
    }
    for _i, _label in enumerate(_sm_labels[:6]):
        _r = 16 + _i
        _lc = ws.cell(row=_r, column=4)
        _lc.value = _label
        _lc.font = Font(bold=True, size=10)
        _lc.fill = PatternFill("solid", fgColor="F5F5F5")
        _lc.alignment = Alignment(horizontal="left", vertical="center")
        _vc = ws.cell(row=_r, column=5)
        _vc.value = _sm_counts.get(_label, 0)
        _vc.fill = PatternFill("solid", fgColor=_sm_palette.get(_label, "F2F2F2"))
        _vc.font = Font(bold=True, size=11)
        _vc.alignment = Alignment(horizontal="center", vertical="center")
        _vc.border = Border(
            left=Side(style="thin", color="D3D3D3"),
            right=Side(style="thin", color="D3D3D3"),
            top=Side(style="thin", color="D3D3D3"),
            bottom=Side(style="thin", color="D3D3D3"),
        )
        ws.row_dimensions[_r].height = 20

    # Low-emphasis classification notes for the less self-explanatory models.
    _sm_notes = [
        "Hybrid: Arc, VMware, Azure Stack, and cross-environment management.",
        "Supporting Services: shared security, operations, governance, and monitoring.",
        "Other: unclassified or third-party / Marketplace types; not necessarily a concern.",
    ]
    for _i, _note in enumerate(_sm_notes):
        _r = 23 + _i
        ws.merge_cells(start_row=_r, start_column=4, end_row=_r, end_column=6)
        _nc = ws.cell(row=_r, column=4)
        _nc.value = _note
        _nc.font = Font(size=8, italic=True, color="888888")
        _nc.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        _nc.fill = PatternFill("solid", fgColor="FAFAFA")
        ws.row_dimensions[_r].height = 22

    # Service Model pie chart (next to the table)
    _sm_n = len(_sm_labels[:6])
    if _sm_n:
        from openpyxl.chart import PieChart, Reference as _SMRef
        from openpyxl.chart.label import DataLabelList
        from openpyxl.chart.series import DataPoint as _SMDP

        _pie = PieChart()
        _pie.title = "Service Model Distribution"
        _pie.height = 8.5
        _pie.width = 12.0
        _pie.style = 10
        _pdata = _SMRef(ws, min_col=5, min_row=16, max_row=15 + _sm_n)
        _pcats = _SMRef(ws, min_col=4, min_row=16, max_row=15 + _sm_n)
        _pie.add_data(_pdata, titles_from_data=False)
        _pie.set_categories(_pcats)
        _pie.dataLabels = DataLabelList()
        _pie.dataLabels.showVal = True
        _sm_pie_palette = ["1F4E79", "2E86AB", "548235", "BF8F00", "C55A11", "7F7F7F"]
        _pser = _pie.series[0]
        for _di in range(_sm_n):
            _dp = _SMDP(idx=_di)
            _dp.graphicalProperties.solidFill = _sm_pie_palette[_di % len(_sm_pie_palette)]
            _pser.dPt.append(_dp)
        ws.add_chart(_pie, "G15")

    # Section 2: Top Resource Types (with dark header)
    ws.merge_cells("D4:E4")
    ws["D4"].value = "TOP RESOURCE TYPES"
    ws["D4"].fill = section_header_fill
    ws["D4"].font = section_header_font
    ws["D4"].alignment = Alignment(horizontal="center", vertical="center")

    for idx, (rtype, count) in enumerate(summary.get("top_resource_types", [])[:10]):
        row = 5 + idx
        display = rtype.split("/")[-1].replace("-", " ").title()
        type_cell = ws.cell(row=row, column=4)
        type_cell.value = display
        type_cell.fill = PatternFill("solid", fgColor="F5F5F5")
        type_cell.font = Font(size=10)

        count_cell = ws.cell(row=row, column=5)
        count_cell.value = count
        count_cell.fill = PatternFill("solid", fgColor="D6E4F0")
        count_cell.font = Font(bold=True, size=10)
        count_cell.alignment = Alignment(horizontal="center")
        count_cell.border = Border(
            left=Side(style="thin", color="D3D3D3"),
            right=Side(style="thin", color="D3D3D3"),
        )

    # Section 3: Advisor by Pillar (with dark header)
    ws.merge_cells("G4:H4")
    ws["G4"].value = "ADVISOR BY WAF PILLAR"
    ws["G4"].fill = section_header_fill
    ws["G4"].font = section_header_font
    ws["G4"].alignment = Alignment(horizontal="center", vertical="center")

    for idx, (pillar, count) in enumerate(summary.get("advisor_by_pillar", {}).items()):
        row = 5 + idx
        pillar_cell = ws.cell(row=row, column=7)
        pillar_cell.value = pillar
        pillar_cell.fill = PatternFill("solid", fgColor="F5F5F5")
        pillar_cell.font = Font(size=10)

        count_cell = ws.cell(row=row, column=8)
        count_cell.value = count
        count_cell.fill = PatternFill("solid", fgColor="D6E4F0")
        count_cell.font = Font(bold=True, size=10)
        count_cell.alignment = Alignment(horizontal="center")
        count_cell.border = Border(
            left=Side(style="thin", color="D3D3D3"),
            right=Side(style="thin", color="D3D3D3"),
        )

    # Section 3c/3d: Modernization Signals + Defender Posture (top-right,
    # in the empty columns J+, aligned with the KPI band — no chart overlap).
    _thin = Border(
        left=Side(style="thin", color="D3D3D3"),
        right=Side(style="thin", color="D3D3D3"),
        top=Side(style="thin", color="D3D3D3"),
        bottom=Side(style="thin", color="D3D3D3"),
    )

    def _mband(_s):
        if _s is None:
            return "9AA0A6"
        if _s <= 33:
            return "C00000"
        if _s <= 66:
            return "C55A11"
        return "2E7D32"

    def _kv(row, lcol, label, value, fill, *, white=False, span_value=0):
        _lc = ws.cell(row=row, column=lcol)
        _lc.value = label
        _lc.font = Font(bold=True, size=10)
        _lc.fill = PatternFill("solid", fgColor="F5F5F5")
        _lc.alignment = Alignment(horizontal="left", vertical="center")
        _vc = ws.cell(row=row, column=lcol + 1)
        _vc.value = value
        _vc.fill = PatternFill("solid", fgColor=fill)
        _vc.font = Font(bold=True, size=10, color="FFFFFF" if white else "000000")
        _vc.alignment = Alignment(horizontal="center", vertical="center")
        _vc.border = _thin
        if span_value:
            ws.merge_cells(start_row=row, start_column=lcol + 1,
                           end_row=row, end_column=lcol + 1 + span_value)
        ws.row_dimensions[row].height = 18

    # --- MODERNIZATION SIGNALS (cols J:K) ---
    from processors.modernization import build_modernization_assessment
    _mod = build_modernization_assessment(scan_data)
    ws.merge_cells("J4:K4")
    ws["J4"].value = "MODERNIZATION SIGNALS"
    ws["J4"].fill = section_header_fill
    ws["J4"].font = section_header_font
    ws["J4"].alignment = Alignment(horizontal="center", vertical="center")
    if _mod.get("available"):
        _ms = _mod.get("summary", {})
        _readiness = _ms.get("readiness")
        _kv(5, 10, "Readiness (0-100)",
            "N/A" if _readiness is None else _readiness, _mband(_readiness), white=True)
        _kv(6, 10, "IaaS Share", f"{_ms.get('iaas_pct', 0)}%", "D6E4F0")
        _kv(7, 10, "PaaS Share", f"{_ms.get('paas_pct', 0)}%", "E2EFDA")
        ws.merge_cells("J8:K8")
        ws["J8"].value = "TOP OPPORTUNITIES"
        ws["J8"].fill = PatternFill("solid", fgColor="1F4E79")
        ws["J8"].font = Font(bold=True, size=9, color="FFFFFF")
        ws["J8"].alignment = Alignment(horizontal="center", vertical="center")
        _tops = _ms.get("top_opportunities", [])[:3]
        for _i, _o in enumerate(_tops):
            _kv(9 + _i, 10, _o["name"],
                "N/A" if _o["score"] is None else _o["score"], _mband(_o["score"]), white=True)
        _note_row = 9 + max(len(_tops), 1)
        ws.merge_cells(start_row=_note_row, start_column=10, end_row=_note_row, end_column=11)
        ws.cell(row=_note_row, column=10).value = "INFERRED — indicative only"
        ws.cell(row=_note_row, column=10).font = Font(size=8, italic=True, color="888888")
    else:
        ws.merge_cells("J5:K5")
        ws["J5"].value = "Not available"
        ws["J5"].font = Font(size=9, italic=True, color="888888")

    # --- DEFENDER FOR CLOUD — PLAN POSTURE (cols M:P) ---
    _posture = scan_data.get("defender_posture", []) or []
    ws.merge_cells("M4:P4")
    ws["M4"].value = "DEFENDER FOR CLOUD — PLAN POSTURE"
    ws["M4"].fill = section_header_fill
    ws["M4"].font = section_header_font
    ws["M4"].alignment = Alignment(horizontal="center", vertical="center")
    if _posture:
        _enabled = sum(1 for p in _posture if p.get("enabled"))
        _total_plans = len(_posture)
        _free = _total_plans - _enabled
        try:
            from collectors.defender_posture import build_servers_resource_coverage
            _srv = build_servers_resource_coverage(_posture, scan_data.get("resources_by_type", {}))
            _srv_total = len(_srv)
            _srv_cov = sum(1 for r in _srv if r.get("covered"))
        except Exception:
            _srv_total = _srv_cov = 0
        _srv_unc = _srv_total - _srv_cov
        _kv(5, 13, "Plans Enabled", f"{_enabled}/{_total_plans}", "E2EFDA", span_value=2)
        _kv(6, 13, "Plans Off", _free, "FCE4D6" if _free else "E2EFDA", span_value=2)
        _kv(7, 13, "Servers Covered", f"{_srv_cov}/{_srv_total}",
            "F8D7DA" if _srv_unc else "E2EFDA", span_value=2)
        _kv(8, 13, "Not Covered", _srv_unc, "F8D7DA" if _srv_unc else "E2EFDA", span_value=2)
        try:
            from collectors.defender_pricing import compute_coverage_gap, total_gap_monthly_cost
            _gap = compute_coverage_gap(scan_data.get("resources_by_type", {}), _posture)
            _gap_units = sum(r["total_units"] for r in _gap)
            _gap_unprot = sum(r["unprotected_units"] for r in _gap)
            _gap_cost = total_gap_monthly_cost(_gap)
            _kv(9, 13, "Coverage Gap (est.)", f"${_gap_cost:,.2f}/mo",
                "FCE4D6" if _gap_cost else "E2EFDA", span_value=2)
            _kv(10, 13, "Units Unprotected", f"{_gap_unprot:,}/{_gap_units:,}", "F5F5F5", span_value=2)
        except Exception:
            pass
        _disabled = sorted({p.get("planDisplayName", "") for p in _posture if not p.get("enabled")})
        ws.cell(row=11, column=13).value = "Disabled Plans"
        ws.cell(row=11, column=13).font = Font(bold=True, size=10)
        ws.cell(row=11, column=13).fill = PatternFill("solid", fgColor="F5F5F5")
        ws.merge_cells(start_row=11, start_column=14, end_row=11, end_column=16)
        _dc = ws.cell(row=11, column=14)
        _dc.value = ", ".join(_disabled[:8]) or "None"
        _dc.font = Font(size=9, color="555555")
        _dc.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        ws.row_dimensions[11].height = 28
    else:
        ws.merge_cells("M5:P5")
        ws["M5"].value = "Plan posture not available — requires Reader RBAC (Microsoft.Security/pricings/read)."
        ws["M5"].font = Font(size=9, italic=True, color="888888")
        ws["M5"].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

    _col_widths(ws, [28, 18, 16, 32, 14, 4, 28, 14,
                     3, 30, 13, 3, 26, 15, 13, 13])


    # Section 3b: Business Pillar summary (ALL pillars) + horizontal bar chart
    _pil_counts: Dict[str, int] = {}
    for _rt, _rows in scan_data.get("resources_by_type", {}).items():
        _p = classify_resource_type(_rt)["business_pillar"]
        _pil_counts[_p] = _pil_counts.get(_p, 0) + len(_rows)
    _pil_sorted = sorted(_pil_counts.items(), key=lambda x: -x[1])
    _pil_n = len(_pil_sorted)

    _pil_header_row = 33
    ws.merge_cells(f"A{_pil_header_row}:H{_pil_header_row}")
    ws[f"A{_pil_header_row}"].value = "BUSINESS PILLAR \u2014 ALL"
    ws[f"A{_pil_header_row}"].fill = section_header_fill
    ws[f"A{_pil_header_row}"].font = section_header_font
    ws[f"A{_pil_header_row}"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[_pil_header_row].height = 22

    _pil_tbl_hdr = _pil_header_row + 1
    ws.cell(row=_pil_tbl_hdr, column=1).value = "Pillar"
    ws.cell(row=_pil_tbl_hdr, column=2).value = "Resources"
    _header_style(ws, _pil_tbl_hdr, 2)
    for _i, (_p, _c) in enumerate(_pil_sorted):
        _r = _pil_tbl_hdr + 1 + _i
        ws.cell(row=_r, column=1).value = _p
        ws.cell(row=_r, column=1).fill = PatternFill("solid", fgColor="F5F5F5")
        ws.cell(row=_r, column=1).font = Font(size=10)
        _pvc = ws.cell(row=_r, column=2)
        _pvc.value = _c
        _pvc.fill = PatternFill("solid", fgColor="D6E4F0")
        _pvc.font = Font(bold=True, size=10)
        _pvc.alignment = Alignment(horizontal="center")

    _pil_bar_h = max(6.0, _pil_n * 0.7)
    if _pil_n:
        from openpyxl.chart import BarChart as _PBC, Reference as _PRef
        _pbar = _PBC()
        _pbar.type = "bar"
        _pbar.grouping = "clustered"
        _pbar.title = "Resources by Business Pillar"
        _pbar.height = _pil_bar_h
        _pbar.width = 16
        _pbar.style = 10
        _pbd = _PRef(ws, min_col=2, min_row=_pil_tbl_hdr, max_row=_pil_tbl_hdr + _pil_n)
        _pbc = _PRef(ws, min_col=1, min_row=_pil_tbl_hdr + 1, max_row=_pil_tbl_hdr + _pil_n)
        _pbar.add_data(_pbd, titles_from_data=True)
        _pbar.set_categories(_pbc)
        _pbar.x_axis.delete = False
        _pbar.x_axis.tickLblPos = "nextTo"
        _pbar.legend = None
        ws.add_chart(_pbar, f"D{_pil_tbl_hdr}")

    _pil_bar_rows = int(_pil_bar_h / 0.5) + 2
    _pil_band_end = _pil_tbl_hdr + max(_pil_n, _pil_bar_rows)

    # Section 4: Microsoft References (5 links)
    _REFERENCES = [
        ("Azure Well-Architected Framework overview",
         "https://learn.microsoft.com/en-us/azure/well-architected/"),
        ("CAF Security design area",
         "https://learn.microsoft.com/en-us/azure/cloud-adoption-framework/ready/landing-zone/design-area/security"),
        ("Azure Advisor documentation",
         "https://learn.microsoft.com/en-us/azure/advisor/advisor-overview"),
        ("Microsoft Defender for Cloud overview",
         "https://learn.microsoft.com/en-us/azure/defender-for-cloud/defender-for-cloud-introduction"),
        ("Azure Policy documentation",
         "https://learn.microsoft.com/en-us/azure/governance/policy/overview"),
    ]
    ref_start_row = _pil_band_end + 2

    ws.merge_cells(f"A{ref_start_row}:H{ref_start_row}")
    ref_header = ws[f"A{ref_start_row}"]
    ref_header.value = "MICROSOFT REFERENCES — CAF & WAF"
    ref_header.fill = section_header_fill
    ref_header.font = section_header_font
    ref_header.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[ref_start_row].height = 24

    for i, (label, url) in enumerate(_REFERENCES):
        ref_row = ref_start_row + 1 + i
        ws.merge_cells(f"A{ref_row}:H{ref_row}")
        c = ws[f"A{ref_row}"]
        c.value = f"{label} — {url}"
        c.font = Font(size=9, color="1F4E79", italic=True)
        c.fill = PatternFill("solid", fgColor="D6E4F0")
        c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        ws.row_dimensions[ref_row].height = 18

    # Section 5: Active Regions Distribution (table + bar chart)
    from openpyxl.chart import BarChart, Reference

    region_data = summary.get("resources_by_location", {})
    top_regions = sorted(region_data.items(), key=lambda x: -x[1])[:12]

    chart_data_start = ref_start_row + len(_REFERENCES) + 2  # gap row

    ws.merge_cells(f"A{chart_data_start}:H{chart_data_start}")
    rgn_header = ws[f"A{chart_data_start}"]
    rgn_header.value = "ACTIVE REGIONS DISTRIBUTION"
    rgn_header.fill = section_header_fill
    rgn_header.font = section_header_font
    rgn_header.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[chart_data_start].height = 24

    # Write data table (cols A-B) and chart data (cols A-B, reused)
    tbl_header_row = chart_data_start + 1
    ws.cell(row=tbl_header_row, column=1).value = "Region"
    ws.cell(row=tbl_header_row, column=2).value = "Resources"
    _header_style(ws, tbl_header_row, 2)
    ws.row_dimensions[tbl_header_row].height = 20

    if top_regions:
        for r_idx, (region, count) in enumerate(top_regions):
            data_row = tbl_header_row + 1 + r_idx
            ws.cell(row=data_row, column=1).value = region
            ws.cell(row=data_row, column=2).value = count
            ws.cell(row=data_row, column=1).fill = PatternFill("solid", fgColor="F5F5F5")
            ws.cell(row=data_row, column=2).fill = PatternFill("solid", fgColor="D6E4F0")
            ws.cell(row=data_row, column=2).alignment = Alignment(horizontal="center")
            ws.row_dimensions[data_row].height = 18

        chart = BarChart()
        chart.type = "bar"
        chart.grouping = "clustered"
        chart.title = "Active Regions — Resource Count"
        chart.y_axis.title = "Resources"
        chart.style = 10
        chart.width = 18
        chart.height = 12

        data_ref = Reference(
            ws,
            min_col=2, max_col=2,
            min_row=tbl_header_row,
            max_row=tbl_header_row + len(top_regions),
        )
        cats_ref = Reference(
            ws,
            min_col=1,
            min_row=tbl_header_row + 1,
            max_row=tbl_header_row + len(top_regions),
        )
        chart.add_data(data_ref, titles_from_data=True)
        chart.set_categories(cats_ref)

        # Ensure the category axis (region names, left side) is visible.
        chart.x_axis.delete = False
        chart.x_axis.tickLblPos = "nextTo"

        # Apply a distinct colour per bar using DataPoint so each region is visually distinct.
        # Palette: 12 colours derived from the Microsoft blue-to-teal/green spectrum.
        _BAR_PALETTE = [
            "1F4E79", "2E75B6", "4472C4", "70AD47",
            "ED7D31", "FFC000", "5BA3C9", "A9D18E",
            "FF7C80", "9DC3E6", "F4B183", "C5E0B4",
        ]
        from openpyxl.chart.data_source import NumDataSource, NumRef
        from openpyxl.chart.series import DataPoint
        ser = chart.series[0]
        for idx in range(len(top_regions)):
            pt = DataPoint(idx=idx)
            pt.graphicalProperties.solidFill = _BAR_PALETTE[idx % len(_BAR_PALETTE)]
            pt.graphicalProperties.line.solidFill = _BAR_PALETTE[idx % len(_BAR_PALETTE)]
            ser.dPt.append(pt)

        # Hide the legend — categories on the axis already identify each bar.
        chart.legend = None

        ws.add_chart(chart, f"C{chart_data_start}")

    # Section 6: Data Collection Notes (subtle — mirrors the HTML reports)
    collection_warnings = scan_data.get("collection_warnings", [])
    notes_start = chart_data_start + max(26, len(top_regions) + 4)
    ws.merge_cells(f"A{notes_start}:H{notes_start}")
    _nh = ws[f"A{notes_start}"]
    _nh.value = "DATA COLLECTION NOTES"
    _nh.fill = section_header_fill
    _nh.font = section_header_font
    _nh.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[notes_start].height = 22

    _note_font = Font(size=9, color="8A8A8A", italic=True)
    _note_fill = PatternFill("solid", fgColor="F7F7F7")
    _note_align = Alignment(horizontal="left", vertical="center", wrap_text=True)
    if collection_warnings:
        for _i, _w in enumerate(collection_warnings):
            _r = notes_start + 1 + _i
            ws.merge_cells(f"A{_r}:H{_r}")
            _c = ws[f"A{_r}"]
            _collector = (_w.get("collector", "") or "").strip()
            _level = (_w.get("level", "") or "").strip()
            _msg = (_w.get("message", "") or "").strip()
            _prefix = f"[{_level}] " if _level else ""
            _c.value = f"{_prefix}{_collector}: {_msg}" if _collector else f"{_prefix}{_msg}"
            _c.font = _note_font
            _c.fill = _note_fill
            _c.alignment = _note_align
            ws.row_dimensions[_r].height = 16
    else:
        _r = notes_start + 1
        ws.merge_cells(f"A{_r}:H{_r}")
        _c = ws[f"A{_r}"]
        _c.value = (
            "No data collection warnings recorded for this scan — "
            "all requested sources returned successfully."
        )
        _c.font = _note_font
        _c.fill = _note_fill
        _c.alignment = _note_align
        ws.row_dimensions[_r].height = 16


def _write_classification_sheet(wb, scan_data: dict, sheet_map: Dict[str, str]):
    """Resource classification taxonomy — Technical Category / Business Pillar /
    Service Model / Publisher, with per-type counts and summary pivots."""
    from openpyxl.styles import Alignment, Font, PatternFill
    from processors.classifier import classify_resource_type, service_model_order

    ws = wb.create_sheet("Classification")
    ws.sheet_view.showGridLines = False

    resources_by_type = scan_data.get("resources_by_type", {})

    by_service: Dict[str, int] = {}
    by_pillar: Dict[str, int] = {}
    by_category: Dict[str, int] = {}
    rows_data = []
    for rtype, resources in resources_by_type.items():
        cls = classify_resource_type(rtype)
        n = len(resources)
        by_service[cls["service_model"]] = by_service.get(cls["service_model"], 0) + n
        by_pillar[cls["business_pillar"]] = by_pillar.get(cls["business_pillar"], 0) + n
        by_category[cls["technical_category"]] = by_category.get(cls["technical_category"], 0) + n
        rows_data.append((rtype, cls, n))

    ws.merge_cells("A1:G1")
    title = ws["A1"]
    title.value = "Resource Classification — Assessment Taxonomy"
    title.font = Font(bold=True, size=16, color="FFFFFF")
    title.fill = PatternFill("solid", fgColor=COLOR_HEADER_BG)
    title.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 30
    ws["A2"] = (
        "Service Model: IaaS / PaaS / SaaS / Hybrid / Supporting Services / Other  ·  "
        "Publisher: Microsoft vs Third-party  ·  editable in config/resource_classification.yaml"
    )
    ws["A2"].font = Font(size=9, italic=True, color="888888")

    section_fill = PatternFill("solid", fgColor="1F4E79")
    section_font = Font(bold=True, size=11, color="FFFFFF")

    # Summary — by Service Model (A4) and by Business Pillar (D4)
    ws.merge_cells("A4:B4")
    ws["A4"].value = "BY SERVICE MODEL"
    ws["A4"].fill = section_fill
    ws["A4"].font = section_font
    ws["A4"].alignment = Alignment(horizontal="center")
    sm_labels = [m for m in service_model_order() if by_service.get(m)]
    sm_labels += [m for m in by_service if m not in sm_labels]
    r = 5
    for label in sm_labels:
        ws.cell(row=r, column=1).value = label
        ws.cell(row=r, column=1).font = Font(bold=True, size=10)
        ws.cell(row=r, column=1).fill = PatternFill("solid", fgColor="F5F5F5")
        vc = ws.cell(row=r, column=2)
        vc.value = by_service[label]
        vc.fill = PatternFill("solid", fgColor="D6E4F0")
        vc.alignment = Alignment(horizontal="center")
        vc.font = Font(bold=True, size=10)
        r += 1

    ws.merge_cells("D4:E4")
    ws["D4"].value = "BY BUSINESS PILLAR"
    ws["D4"].fill = section_fill
    ws["D4"].font = section_font
    ws["D4"].alignment = Alignment(horizontal="center")
    r = 5
    for label, cnt in sorted(by_pillar.items(), key=lambda x: -x[1]):
        ws.cell(row=r, column=4).value = label
        ws.cell(row=r, column=4).font = Font(bold=True, size=10)
        ws.cell(row=r, column=4).fill = PatternFill("solid", fgColor="F5F5F5")
        vc = ws.cell(row=r, column=5)
        vc.value = cnt
        vc.fill = PatternFill("solid", fgColor="D6E4F0")
        vc.alignment = Alignment(horizontal="center")
        vc.font = Font(bold=True, size=10)
        r += 1

    ws.merge_cells("G4:H4")
    ws["G4"].value = "BY TECHNICAL CATEGORY"
    ws["G4"].fill = section_fill
    ws["G4"].font = section_font
    ws["G4"].alignment = Alignment(horizontal="center")
    r = 5
    for label, cnt in sorted(by_category.items(), key=lambda x: -x[1]):
        ws.cell(row=r, column=7).value = label
        ws.cell(row=r, column=7).font = Font(bold=True, size=10)
        ws.cell(row=r, column=7).fill = PatternFill("solid", fgColor="F5F5F5")
        vc = ws.cell(row=r, column=8)
        vc.value = cnt
        vc.fill = PatternFill("solid", fgColor="D6E4F0")
        vc.alignment = Alignment(horizontal="center")
        vc.font = Font(bold=True, size=10)
        r += 1

    # Detail table (one row per resource type present in the scan)
    table_header_row = max(5 + len(sm_labels), 5 + len(by_pillar), 5 + len(by_category)) + 1
    headers = [
        "Resource Type", "Technical Category", "Business Pillar",
        "Service Model", "Publisher", "Count", "Sheet Tab",
    ]
    for c, h in enumerate(headers, 1):
        ws.cell(row=table_header_row, column=c).value = h
    _header_style(ws, table_header_row, len(headers))
    _auto_filter(ws, table_header_row, len(headers))
    _freeze(ws, table_header_row + 1)

    rows_data.sort(key=lambda item: (item[1]["business_pillar"], -item[2]))
    rr = table_header_row + 1
    for rtype, cls, n in rows_data:
        ws.cell(row=rr, column=1).value = rtype
        ws.cell(row=rr, column=2).value = cls["technical_category"]
        ws.cell(row=rr, column=3).value = cls["business_pillar"]
        ws.cell(row=rr, column=4).value = cls["service_model"]
        ws.cell(row=rr, column=5).value = cls["publisher"]
        ws.cell(row=rr, column=6).value = n
        ws.cell(row=rr, column=7).value = sheet_map.get(rtype, "(AllResources)")
        rr += 1

    _col_widths(ws, [46, 28, 20, 20, 14, 10, 34, 12])


_FAM_ACRONYMS = {"aks", "ai", "sql", "vm", "api", "aci", "acr", "cdn", "waf", "dns", "iot"}


_FAM_LABEL_OVERRIDES = {"openai_ai_services": "OpenAI / AI Services"}


def _fam_label(key: str) -> str:
    if key in _FAM_LABEL_OVERRIDES:
        return _FAM_LABEL_OVERRIDES[key]
    return " ".join(w.upper() if w.lower() in _FAM_ACRONYMS else w.capitalize()
                    for w in str(key).split("_"))


def _mod_evidence_text(d: dict) -> str:
    ev = d.get("evidence", {}) or {}
    m = d.get("method")
    if m == "proportion":
        return f"modern={ev.get('modern', 0)}, legacy={ev.get('legacy', 0)}, total={ev.get('total', 0)}"
    if m == "presence":
        fams = ev.get("families", {}) or {}
        present = [f"{_fam_label(k)}={v}" for k, v in fams.items() if v]
        absent = [_fam_label(k) for k, v in fams.items() if not v]
        base = f"{ev.get('present', 0)}/{ev.get('total', 0)} families"
        if present:
            base += " present: " + ", ".join(present)
        if absent:
            base += " · absent: " + ", ".join(absent)
        return base
    if m == "security":
        cov = ev.get("defender_coverage_pct")
        return f"Defender {cov if cov is not None else 'n/a'}% · High misconfig {ev.get('high_misconfig', 0)}"
    if m == "governance":
        return (f"tags {ev.get('tag_pct')}% · compliance {ev.get('compliance_pct')}% · "
                f"MGs {ev.get('mg_count', 0)}")
    if m == "footprint":
        return (f"Microsoft {ev.get('microsoft', 0)} · Third-party {ev.get('third_party', 0)} "
                f"({ev.get('third_party_pct', 0)}%)")
    return ""


def _write_modernization_sheet(wb, scan_data: dict):
    """Cloud Modernization & Opportunity — per-dimension inferred maturity/adoption
    scores (0-100) with confidence, evidence, and deterministic narrative."""
    from openpyxl.styles import Alignment, Font, PatternFill
    from processors.modernization import build_modernization_assessment

    a = build_modernization_assessment(scan_data)
    if not a.get("available"):
        return

    ws = wb.create_sheet("ModernizationSignals")
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:K1")
    title = ws["A1"]
    title.value = "Cloud Modernization & Opportunity — Signals (INFERRED)"
    title.font = Font(bold=True, size=16, color="FFFFFF")
    title.fill = PatternFill("solid", fgColor=COLOR_HEADER_BG)
    title.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 30

    ws["A2"] = a.get("inferred_label", "")
    ws["A2"].font = Font(size=9, italic=True, color="888888")
    summary = a.get("summary", {})
    ws["A3"] = summary.get("narrative", "")
    ws["A3"].font = Font(size=10, bold=True, color="1F4E79")
    ws.merge_cells("A3:K3")
    ws.row_dimensions[3].height = 28
    ws["A3"].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

    # ---- Part 1: Current environment (As-Is) mini-tables ------------------
    section_fill = PatternFill("solid", fgColor="1F4E79")
    section_font = Font(bold=True, size=11, color="FFFFFF")
    asis = summary.get("as_is", {})
    sm = asis.get("service_model", {})
    bp = asis.get("business_pillar", {})
    tc = asis.get("technical_category", {})
    sm_counts, sm_pct = sm.get("counts", {}), sm.get("pct", {})
    bp_counts, bp_pct = bp.get("counts", {}), bp.get("pct", {})
    tc_counts, tc_pct = tc.get("counts", {}), tc.get("pct", {})

    ws.merge_cells("A5:C5")
    ws["A5"].value = "CURRENT ENVIRONMENT — BY SERVICE MODEL"
    ws["A5"].fill = section_fill
    ws["A5"].font = section_font
    ws["A5"].alignment = Alignment(horizontal="center")
    ws.merge_cells("E5:G5")
    ws["E5"].value = "BY BUSINESS PILLAR"
    ws["E5"].fill = section_fill
    ws["E5"].font = section_font
    ws["E5"].alignment = Alignment(horizontal="center")
    ws.merge_cells("I5:K5")
    ws["I5"].value = "BY TECHNICAL CATEGORY"
    ws["I5"].fill = section_fill
    ws["I5"].font = section_font
    ws["I5"].alignment = Alignment(horizontal="center")

    def _asis_rows(col0, counts, pct):
        r = 6
        for label, cnt in sorted(counts.items(), key=lambda x: -x[1]):
            lc = ws.cell(row=r, column=col0)
            lc.value = label
            lc.font = Font(bold=True, size=10)
            lc.fill = PatternFill("solid", fgColor="F5F5F5")
            vc = ws.cell(row=r, column=col0 + 1)
            vc.value = cnt
            vc.fill = PatternFill("solid", fgColor="D6E4F0")
            vc.alignment = Alignment(horizontal="center")
            vc.font = Font(bold=True, size=10)
            pc = ws.cell(row=r, column=col0 + 2)
            pc.value = f"{pct.get(label, 0)}%"
            pc.fill = PatternFill("solid", fgColor="EAF1FB")
            pc.alignment = Alignment(horizontal="center")
            pc.font = Font(size=10, color="1F4E79")
            r += 1
        return r

    end_sm = _asis_rows(1, sm_counts, sm_pct)
    end_bp = _asis_rows(5, bp_counts, bp_pct)
    end_tc = _asis_rows(9, tc_counts, tc_pct)
    ctx_row = max(end_sm, end_bp, end_tc)
    ws.cell(row=ctx_row, column=1).value = (
        f"Context: {asis.get('third_party_pct', 0)}% of resources are third-party / Marketplace "
        f"publishers (neutral — informational only)."
    )
    ws.cell(row=ctx_row, column=1).font = Font(size=9, italic=True, color="888888")
    ws.merge_cells(start_row=ctx_row, start_column=1, end_row=ctx_row, end_column=11)

    header_row = ctx_row + 2
    headers = ["Dimension", "Score", "Level", "Confidence", "Opportunity",
               "Signal (inferred)", "Evidence", "Frameworks"]
    ws.cell(row=header_row - 1, column=1).value = "MODERNIZATION SIGNALS BY DIMENSION"
    ws.cell(row=header_row - 1, column=1).font = section_font
    ws.cell(row=header_row - 1, column=1).fill = section_fill
    ws.merge_cells(start_row=header_row - 1, start_column=1, end_row=header_row - 1, end_column=8)
    ws.cell(row=header_row - 1, column=1).alignment = Alignment(horizontal="center")
    for c, h in enumerate(headers, 1):
        ws.cell(row=header_row, column=c).value = h
    _header_style(ws, header_row, len(headers))
    _auto_filter(ws, header_row, len(headers))
    _freeze(ws, header_row + 1)

    rr = header_row + 1
    for d in a.get("dimensions", []):
        ws.cell(row=rr, column=1).value = d["name"]
        ws.cell(row=rr, column=1).font = Font(bold=True, size=10)

        score_cell = ws.cell(row=rr, column=2)
        score_cell.value = "N/A" if d["score"] is None else d["score"]
        score_cell.alignment = Alignment(horizontal="center")
        score_cell.font = Font(bold=True, color="FFFFFF")
        score_cell.fill = PatternFill("solid", fgColor=(d.get("color") or "#9AA0A6").lstrip("#"))

        ws.cell(row=rr, column=3).value = d["level"]
        conf_cell = ws.cell(row=rr, column=4)
        conf_cell.value = d["confidence"]
        if d["confidence"] == "Low":
            conf_cell.font = Font(color="999999", italic=True)
        ws.cell(row=rr, column=5).value = "YES" if d["opportunity"] else "—"
        if d["opportunity"]:
            ws.cell(row=rr, column=5).font = Font(bold=True, color="C00000")
        ws.cell(row=rr, column=6).value = d["narrative"]
        ws.cell(row=rr, column=7).value = _mod_evidence_text(d)
        ws.cell(row=rr, column=8).value = "; ".join(
            f.get("name", "") for f in d.get("framework_refs", []))
        rr += 1

    # Top opportunities recap
    tops = summary.get("top_opportunities", [])
    if tops:
        rr += 1
        ws.cell(row=rr, column=1).value = "TOP OPPORTUNITIES"
        ws.cell(row=rr, column=1).font = Font(bold=True, size=11, color="FFFFFF")
        ws.cell(row=rr, column=1).fill = PatternFill("solid", fgColor="1F4E79")
        rr += 1
        for o in tops:
            ws.cell(row=rr, column=1).value = o["name"]
            ws.cell(row=rr, column=2).value = o["score"]
            ws.cell(row=rr, column=2).alignment = Alignment(horizontal="center")
            ws.cell(row=rr, column=3).value = f"confidence: {o['confidence']}"
            rr += 1

    # ---- Legend / how to read -------------------------------------------
    rr += 1
    ws.cell(row=rr, column=1).value = "HOW TO READ"
    ws.cell(row=rr, column=1).font = Font(bold=True, size=11, color="FFFFFF")
    ws.cell(row=rr, column=1).fill = section_fill
    ws.merge_cells(start_row=rr, start_column=1, end_row=rr, end_column=8)
    ws.cell(row=rr, column=1).alignment = Alignment(horizontal="center")
    rr += 1
    legend = [
        ("Score 0–33", "High opportunity", "C00000"),
        ("Score 34–66", "Intermediate", "C55A11"),
        ("Score 67–100", "Mature adoption", "2E7D32"),
        ("Confidence", "High / Medium / Low — Low = weak or partial evidence (read as directional)", None),
        ("Opportunity = YES", "score is below this dimension's modernization threshold", None),
        ("Adoption breadth", "presence dimensions measure how many service families exist, not depth/volume", None),
        ("Context rows", "neutral, informational signals (e.g. footprint) — not an opportunity", None),
    ]
    for label, desc, color in legend:
        lc = ws.cell(row=rr, column=1)
        lc.value = label
        lc.font = Font(bold=True, size=10, color="FFFFFF" if color else "1F4E79")
        if color:
            lc.fill = PatternFill("solid", fgColor=color)
        dc = ws.cell(row=rr, column=2)
        dc.value = desc
        dc.font = Font(size=9, color="555555")
        ws.merge_cells(start_row=rr, start_column=2, end_row=rr, end_column=8)
        rr += 1

    _col_widths(ws, [34, 10, 14, 12, 12, 58, 46, 44])


def _write_subscriptions_sheet(wb, scan_data: dict):
    ws = wb.create_sheet("Subscriptions")
    headers = [
        "Subscription",
        "Subscription ID",
        "State",
        "Tenant ID",
        "Resource Group",
        "Location",
        "Resource Type",
        "Resources Count",
    ]
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col).value = h
    _header_style(ws, 1, len(headers))
    _auto_filter(ws, 1, len(headers))
    _freeze(ws, 2)

    subscriptions = {
        sub.get("subscriptionId", ""): sub
        for sub in scan_data.get("subscriptions", [])
    }
    grouped: Dict[tuple, int] = {}
    for rtype, resources in scan_data.get("resources_by_type", {}).items():
        for resource in resources:
            sid = resource.get("subscriptionId", "")
            key = (
                sid,
                resource.get("resourceGroup", ""),
                resource.get("location", ""),
                rtype,
            )
            grouped[key] = grouped.get(key, 0) + 1

    row = 2
    for (sid, resource_group, location, rtype), count in sorted(grouped.items()):
        sub = subscriptions.get(sid, {})
        ws.cell(row=row, column=1).value = sub.get("displayName", "")
        ws.cell(row=row, column=2).value = sid
        ws.cell(row=row, column=3).value = sub.get("state", "")
        ws.cell(row=row, column=4).value = sub.get("tenantId", "")
        ws.cell(row=row, column=5).value = resource_group
        ws.cell(row=row, column=6).value = location
        ws.cell(row=row, column=7).value = rtype
        ws.cell(row=row, column=8).value = count
        row += 1

    if row == 2:
        for sub in scan_data.get("subscriptions", []):
            ws.cell(row=row, column=1).value = sub.get("displayName", "")
            ws.cell(row=row, column=2).value = sub.get("subscriptionId", "")
            ws.cell(row=row, column=3).value = sub.get("state", "")
            ws.cell(row=row, column=4).value = sub.get("tenantId", "")
            ws.cell(row=row, column=8).value = 0
            row += 1

    _col_widths(ws, [36, 38, 12, 38, 32, 18, 46, 16])


def _write_all_resources_sheet(wb, scan_data: dict, include_tags: bool, sheet_map: Dict[str, str]):
    from collectors.resources import load_enrichment_config
    from processors.classifier import classify_resource_type
    from processors.normalizer import safe_str

    ws = wb.create_sheet("AllResources")

    resources_by_type = scan_data.get("resources_by_type", {})
    inventory_config = load_enrichment_config()
    sub_names = _subscription_name_map(scan_data)

    headers = [
        "Resource Tab",
        "Name", "Display Type", "Resource Type", "Category", "Business Pillar", "Service Model",
        "Location", "Resource Group",
        "Subscription", "Subscription ID", "Tenant ID",
        "Kind", "SKU Name", "SKU Tier", "SKU Size", "SKU Family",
        "Provisioning State", "Created Time", "Identity Type", "Zones",
    ]
    if include_tags:
        headers.append("Tags")
    headers.extend(["Resource ID", "Raw Properties"])

    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col).value = h
    _header_style(ws, 1, len(headers))
    _auto_filter(ws, 1, len(headers))
    _freeze(ws, 2)

    row = 2
    for rtype, resources in resources_by_type.items():
        tab_name = sheet_map.get(rtype, "(AllResources)")
        _cls = classify_resource_type(rtype)
        for resource in resources:
            sid = resource.get("subscriptionId", "")

            col = 1
            ws.cell(row=row, column=col).value = tab_name
            col += 1
            for val in [
                resource.get("name", ""),
                _resource_display_name(rtype, inventory_config),
                rtype,
                _resource_category(rtype),
                _cls["business_pillar"],
                _cls["service_model"],
                resource.get("location", ""),
                resource.get("resourceGroup", ""),
                sub_names.get(sid, ""),
                sid,
                resource.get("tenantId", ""),
                safe_str(resource.get("kind", "")),
                _sku_value(resource, "name"),
                _sku_value(resource, "tier"),
                _sku_value(resource, "size"),
                _sku_value(resource, "family"),
                safe_str(_common_resource_value(resource, "provisioningState")),
                safe_str(_common_resource_value(resource, "createdTime")),
                safe_str(_common_resource_value(resource, "identityType")),
                safe_str(resource.get("zones", "")),
            ]:
                ws.cell(row=row, column=col).value = val
                col += 1

            if include_tags:
                tags = resource.get("tags") or {}
                ws.cell(row=row, column=col).value = json.dumps(tags) if tags else ""
                col += 1

            ws.cell(row=row, column=col).value = resource.get("id", "")
            col += 1
            ws.cell(row=row, column=col).value = safe_str(resource.get("properties", ""))
            row += 1

    _col_widths(ws, [22, 35, 28, 46, 18, 22, 18, 20, 30, 36, 38, 38, 18, 24, 18, 16, 18, 22, 22, 18, 20, 60, 60, 80])


def _write_resource_type_sheets(wb, scan_data: dict, include_tags: bool, sheet_map: Dict[str, str]):
    from collectors.resources import load_enrichment_config
    from processors.normalizer import (
        safe_str,
    )

    resources_by_type = scan_data.get("resources_by_type", {})
    inventory_config = load_enrichment_config()
    sub_names = _subscription_name_map(scan_data)

    # Build security findings index: resource_id_lower -> [findings]
    findings_index: dict = {}
    for _f in scan_data.get("misconfig_findings", []):
        _rid = _f.get("resourceId", "").lower()
        if _rid:
            findings_index.setdefault(_rid, []).append(_f)

    for rtype, sheet_name in sheet_map.items():
        resources = resources_by_type.get(rtype)
        if not resources:
            continue

        enriched_cols = _enriched_columns(rtype, resources, inventory_config)

        base_headers = [
            "Name", "Resource Group", "Subscription", "Subscription ID",
            "Location", "Kind", "SKU Name", "SKU Tier",
            "Provisioning State", "Created Time", "Identity Type", "Zones",
        ]
        enriched_display = [c.replace("enriched_", "").replace("_", " ").title() for c in enriched_cols]
        extra = ["Tags", "Resource ID", "Raw Properties"] if include_tags else ["Resource ID", "Raw Properties"]
        all_headers = base_headers + enriched_display + extra

        ws = wb.create_sheet(sheet_name)
        for col, h in enumerate(all_headers, 1):
            ws.cell(row=1, column=col).value = h
        _header_style(ws, 1, len(all_headers))
        _auto_filter(ws, 1, len(all_headers))
        _freeze(ws, 2)

        for row, resource in enumerate(resources, 2):
            sid = resource.get("subscriptionId", "")

            ws.cell(row=row, column=1).value = resource.get("name", "")
            ws.cell(row=row, column=2).value = resource.get("resourceGroup", "")
            ws.cell(row=row, column=3).value = sub_names.get(sid, "")
            ws.cell(row=row, column=4).value = sid
            ws.cell(row=row, column=5).value = resource.get("location", "")
            ws.cell(row=row, column=6).value = safe_str(resource.get("kind", ""))
            ws.cell(row=row, column=7).value = _sku_value(resource, "name")
            ws.cell(row=row, column=8).value = _sku_value(resource, "tier")
            ws.cell(row=row, column=9).value = safe_str(_common_resource_value(resource, "provisioningState"))
            ws.cell(row=row, column=10).value = safe_str(_common_resource_value(resource, "createdTime"))
            ws.cell(row=row, column=11).value = safe_str(_common_resource_value(resource, "identityType"))
            ws.cell(row=row, column=12).value = safe_str(resource.get("zones", ""))

            for e_idx, e_col in enumerate(enriched_cols, 13):
                ws.cell(row=row, column=e_idx).value = safe_str(resource.get(e_col))

            next_col = 13 + len(enriched_cols)
            if include_tags:
                tags = resource.get("tags") or {}
                ws.cell(row=row, column=next_col).value = json.dumps(tags) if tags else ""
                next_col += 1
            ws.cell(row=row, column=next_col).value = resource.get("id", "")
            next_col += 1
            ws.cell(row=row, column=next_col).value = safe_str(resource.get("properties", ""))

            # Highlight rows that have security findings
            rid_lower = resource.get("id", "").lower()
            row_findings = findings_index.get(rid_lower, [])
            if row_findings:
                from openpyxl.styles import PatternFill
                from openpyxl.comments import Comment
                warn_fill = PatternFill("solid", fgColor="FFF2CC")
                for c_idx in range(1, len(all_headers) + 1):
                    ws.cell(row=row, column=c_idx).fill = warn_fill
                finding_lines = "\n".join(
                    f"⚠ {f['title']} [{f.get('severity','')}]"
                    for f in row_findings[:5]
                )
                if len(row_findings) > 5:
                    finding_lines += f"\n... +{len(row_findings) - 5} more findings"
                cmt = Comment(finding_lines, "ATI Security")
                cmt.width = 320
                cmt.height = max(60, 22 * min(len(row_findings), 5))
                ws.cell(row=row, column=1).comment = cmt

        _col_widths(ws, [34, 28, 34, 38, 18, 18, 22, 16, 20, 22, 18, 16] + [24] * len(enriched_cols) + [40, 60, 80])


def _write_index_sheet(wb):
    """Create a navigation Index sheet listing every tab with a hyperlink,
    positioned right after Overview."""
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.worksheet.hyperlink import Hyperlink

    ws = wb.create_sheet("Index")
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:C1")
    title = ws["A1"]
    title.value = "Index — Workbook Navigation"
    title.font = Font(bold=True, size=16, color="FFFFFF")
    title.fill = PatternFill("solid", fgColor=COLOR_HEADER_BG)
    title.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 32

    ws["A2"].value = "#"
    ws["B2"].value = "Sheet"
    ws["C2"].value = "Go"
    _header_style(ws, 2, 3)
    _freeze(ws, 3)

    link_font = Font(color="0563C1", underline="single")
    row = 3
    n = 1
    for name in wb.sheetnames:
        if name in ("Overview", "Index"):
            continue
        ws.cell(row=row, column=1).value = n
        name_cell = ws.cell(row=row, column=2)
        name_cell.value = name
        name_cell.font = Font(size=10)
        link_cell = ws.cell(row=row, column=3)
        link_cell.value = "Open →"
        link_cell.hyperlink = Hyperlink(
            ref=link_cell.coordinate, location=f"'{name}'!A1", display=name
        )
        link_cell.font = link_font
        ws.row_dimensions[row].height = 18
        row += 1
        n += 1

    _col_widths(ws, [6, 42, 14])

    # Move Index to position 1 (right after Overview at position 0).
    wb.move_sheet(ws, offset=1 - wb.index(ws))


def _add_back_links(wb):
    """Add a '↩ Index' hyperlink on every sheet (except Overview/Index) so users
    can jump back to the navigation Index."""
    from openpyxl.styles import Font
    from openpyxl.worksheet.hyperlink import Hyperlink

    back_font = Font(bold=True, color="0563C1", underline="single")
    for ws in wb.worksheets:
        if ws.title in ("Overview", "Index"):
            continue
        col = ws.max_column + 2
        cell = ws.cell(row=1, column=col)
        cell.value = "↩ Index"
        cell.hyperlink = Hyperlink(
            ref=cell.coordinate, location="'Index'!A1", display="Index"
        )
        cell.font = back_font


def _write_advisor_sheet(wb, scan_data: dict):
    from openpyxl.styles import PatternFill

    ws = wb.create_sheet("AdvisorFindings")
    headers = [
        "WAF Pillar", "Category", "Impact", "Impacted Field", "Impacted Value",
        "Short Description", "Solution", "Potential Benefits",
        "Subscription ID", "Resource Group", "Learn More",
    ]
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col).value = h
    _header_style(ws, 1, len(headers))
    _auto_filter(ws, 1, len(headers))
    _freeze(ws, 2)

    IMPACT_COLORS = {"High": COLOR_CRITICAL, "Medium": COLOR_MEDIUM, "Low": COLOR_LOW}

    for row, rec in enumerate(scan_data.get("advisor_data", []), 2):
        impact = rec.get("impact", "")
        vals = [
            rec.get("wafPillar", ""),
            rec.get("category", ""),
            impact,
            rec.get("impactedField", ""),
            rec.get("impactedValue", ""),
            rec.get("shortDescription", ""),
            rec.get("solution", ""),
            rec.get("potentialBenefits", ""),
            rec.get("subscriptionId", ""),
            rec.get("resourceGroup", ""),
            rec.get("learnMoreLink", ""),
        ]
        for col, val in enumerate(vals, 1):
            ws.cell(row=row, column=col).value = val
        if impact in IMPACT_COLORS:
            ws.cell(row=row, column=3).fill = PatternFill("solid", fgColor=IMPACT_COLORS[impact])

    _col_widths(ws, [22, 22, 10, 35, 40, 60, 60, 30, 38, 25, 50])


def _write_policy_sheet(wb, scan_data: dict):
    ws = wb.create_sheet("PolicyCompliance")
    headers = [
        "Resource ID", "Resource Type", "Policy Assignment",
        "Policy Definition", "Effect", "Category",
        "Subscription ID", "Resource Group", "Compliance State",
    ]
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col).value = h
    _header_style(ws, 1, len(headers))
    _auto_filter(ws, 1, len(headers))
    _freeze(ws, 2)

    for row, state in enumerate(scan_data.get("policy_data", []), 2):
        for col, key in enumerate(
            ["resourceId", "resourceType", "policyAssignmentName",
             "policyDefinitionName", "policyDefinitionAction",
             "policyDefinitionCategory", "subscriptionId",
             "resourceGroup", "complianceState"],
            1,
        ):
            ws.cell(row=row, column=col).value = state.get(key, "")

    _col_widths(ws, [55, 40, 35, 35, 12, 25, 38, 28, 16])


def _write_health_sheet(wb, scan_data: dict):
    ws = wb.create_sheet("ResourceHealth")
    headers = [
        "Resource Name", "Availability State", "Summary",
        "Reason Type", "Reason Chronicity", "Occurred Time",
        "Subscription ID", "Resource Group", "Location",
    ]
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col).value = h
    _header_style(ws, 1, len(headers))
    _auto_filter(ws, 1, len(headers))
    _freeze(ws, 2)

    for row, event in enumerate(scan_data.get("health_data", []), 2):
        for col, key in enumerate(
            ["name", "availabilityState", "summary", "reasonType",
             "reasonChronicity", "occurredTime",
             "subscriptionId", "resourceGroup", "location"],
            1,
        ):
            ws.cell(row=row, column=col).value = event.get(key, "")

    _col_widths(ws, [35, 18, 60, 25, 20, 22, 38, 28, 18])


def _write_deprecated_sheet(wb, scan_data: dict):
    from openpyxl.styles import PatternFill

    ws = wb.create_sheet("DeprecatedResources")
    headers = [
        "Resource Name", "Resource Type", "Retirement Date", "Severity",
        "Subscription ID", "Resource Group", "Location",
        "Migration Path", "Announcement URL", "Notes",
    ]
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col).value = h
    _header_style(ws, 1, len(headers))
    _auto_filter(ws, 1, len(headers))
    _freeze(ws, 2)

    for row, match in enumerate(scan_data.get("deprecated_matches", []), 2):
        severity = match.get("severity", "")
        for col, key in enumerate(
            ["resourceName", "resourceType", "retirementDate", "severity",
             "subscriptionId", "resourceGroup", "location",
             "migrationPath", "retirementAnnouncementUrl", "notes"],
            1,
        ):
            ws.cell(row=row, column=col).value = match.get(key, "")
        fill = _severity_fill(severity)
        if fill:
            ws.cell(row=row, column=4).fill = fill

    _col_widths(ws, [35, 40, 16, 12, 38, 28, 18, 55, 55, 50])


def _write_misconfig_sheet(wb, scan_data: dict):
    ws = wb.create_sheet("MisconfigFindings")
    headers = [
        "Rule ID", "Title", "Severity", "WAF Pillar",
        "Resource Name", "Resource Type",
        "Subscription ID", "Resource Group", "Location",
        "Actual Value", "Expected Value",
        "Description", "Documentation URL",
    ]
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col).value = h
    _header_style(ws, 1, len(headers))
    _auto_filter(ws, 1, len(headers))
    _freeze(ws, 2)

    for row, finding in enumerate(scan_data.get("misconfig_findings", []), 2):
        severity = finding.get("severity", "")
        keys = [
            "ruleId", "title", "severity", "wafPillar",
            "resourceName", "resourceType",
            "subscriptionId", "resourceGroup", "location",
            "actualValue", "expectedValue",
            "description", "documentationUrl",
        ]
        for col, key in enumerate(keys, 1):
            ws.cell(row=row, column=col).value = finding.get(key, "")
        fill = _severity_fill(severity)
        if fill:
            ws.cell(row=row, column=3).fill = fill

    _col_widths(ws, [12, 45, 10, 22, 35, 40, 38, 28, 18, 20, 20, 60, 55])


def _write_defender_sheet(wb, scan_data: dict):
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    ws = wb.create_sheet("SecurityAssessments")

    # Reference row — CAF Security Design (row 1)
    _CAF_SECURITY_URL = (
        "https://learn.microsoft.com/en-us/azure/cloud-adoption-framework/"
        "ready/landing-zone/design-area/security"
    )
    ws.merge_cells(f"A1:{get_column_letter(10)}1")
    ref_cell = ws["A1"]
    ref_cell.value = (
        "Reference: Security design in Azure — CAF Landing Zone | Microsoft Learn — "
        + _CAF_SECURITY_URL
    )
    ref_cell.font = Font(italic=True, size=9, color="1F4E79")
    ref_cell.fill = PatternFill("solid", fgColor="D6E4F0")
    ref_cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 20

    headers = [
        "Display Name", "Severity", "Status", "Category",
        "Implementation Effort", "Resource",
        "Subscription ID", "Resource Group",
        "Status Cause", "Remediation",
    ]
    for col, h in enumerate(headers, 1):
        ws.cell(row=2, column=col).value = h
    _header_style(ws, 2, len(headers))
    _auto_filter(ws, 2, len(headers))
    _freeze(ws, 3)

    for row, assessment in enumerate(scan_data.get("defender_data", []), 3):
        severity = assessment.get("severity", "")
        keys = [
            "displayName", "severity", "statusCode", "category",
            "implementationEffort", "resourceDetails",
            "subscriptionId", "resourceGroup",
            "statusCause", "remediationDescription",
        ]
        for col, key in enumerate(keys, 1):
            ws.cell(row=row, column=col).value = assessment.get(key, "")
        fill = _severity_fill(severity)
        if fill:
            ws.cell(row=row, column=2).fill = fill

    _col_widths(ws, [55, 10, 14, 25, 22, 55, 38, 28, 20, 60])


def _write_costs_sheet(wb, scan_data: dict):
    ws = wb.create_sheet("Costs")
    headers = ["Subscription ID", "Resource Group", "Service Name", "Currency", "Total Cost (MTD)"]
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col).value = h
    _header_style(ws, 1, len(headers))
    _auto_filter(ws, 1, len(headers))
    _freeze(ws, 2)

    for row, cost in enumerate(scan_data.get("costs_data", []), 2):
        ws.cell(row=row, column=1).value = cost.get("subscriptionId", "")
        ws.cell(row=row, column=2).value = cost.get("ResourceGroupName", cost.get("resourceGroupName", ""))
        ws.cell(row=row, column=3).value = cost.get("ServiceName", cost.get("serviceName", ""))
        ws.cell(row=row, column=4).value = cost.get("Currency", cost.get("currency", ""))
        ws.cell(row=row, column=5).value = cost.get("Cost", cost.get("totalCost", 0))

    _col_widths(ws, [38, 35, 40, 12, 18])


def _write_defender_cost_estimate_sheet(wb, scan_data: dict):
    """Defender for Cloud cost estimate sheet (resource counts x public pricing)."""
    from openpyxl.styles import Alignment, Font, PatternFill, numbers
    from openpyxl.utils import get_column_letter

    ws = wb.create_sheet("DefenderCostEstimate")

    # Title row
    ws.merge_cells(f"A1:{get_column_letter(8)}1")
    title_cell = ws["A1"]
    title_cell.value = "Defender for Cloud — Cost Estimate (Public List Pricing)"
    title_cell.font = Font(bold=True, size=13, color="FFFFFF")
    title_cell.fill = PatternFill("solid", fgColor="1F4E79")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 26

    # Disclaimer row
    ws.merge_cells(f"A2:{get_column_letter(8)}2")
    disc_cell = ws["A2"]
    disc_cell.value = (
        "NOTE: Estimates based on public list pricing. EA/MCA/CSP discounts, free tiers, and included units are NOT reflected. "
        "Billable unit counts are inferred from Resource Graph data. Actual costs may differ."
    )
    disc_cell.font = Font(italic=True, size=9, color="7F7F7F")
    disc_cell.fill = PatternFill("solid", fgColor="FFF2CC")
    disc_cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws.row_dimensions[2].height = 30

    headers = [
        "Defender Plan", "Billable Unit", "Unit Count",
        "P1 Price/Unit/Mo (USD)", "P1 Est. Monthly Cost",
        "P2 Price/Unit/Mo (USD)", "P2 Est. Monthly Cost",
        "Notes",
    ]
    for col, h in enumerate(headers, 1):
        ws.cell(row=3, column=col).value = h
    _header_style(ws, 3, len(headers))
    _auto_filter(ws, 3, len(headers))
    _freeze(ws, 4)

    # Compute rows using the same logic as HTML report
    resources_by_type = scan_data.get("resources_by_type", {})
    vms = (
        len(resources_by_type.get("microsoft.compute/virtualmachines", []))
        + len(resources_by_type.get("microsoft.hybridcompute/machines", []))
    )
    storage = len(resources_by_type.get("microsoft.storage/storageaccounts", []))
    aks = len(resources_by_type.get("microsoft.containerservice/managedclusters", []))
    app_services = len([
        r for r in resources_by_type.get("microsoft.web/sites", [])
        if "function" not in r.get("kind", "").lower()
    ])
    key_vaults = len(resources_by_type.get("microsoft.keyvault/vaults", []))
    sql = (
        len(resources_by_type.get("microsoft.sql/servers", []))
        + len(resources_by_type.get("microsoft.sql/managedinstances", []))
    )

    plan_rows = [
        ("Defender for Servers", "VM / Arc Machine", vms, 4.95, vms * 4.95, 13.95, vms * 13.95,
         "P1: foundational + Defender XDR. P2: full vuln mgmt + JIT access."),
        ("Defender for Storage", "Storage Account", storage, 10.00, storage * 10.00, None, None,
         "Malware scanning + sensitive data discovery."),
        ("Defender for Containers", "AKS Cluster (est.)", aks, 7.00, aks * 7.00, None, None,
         "$7/vCore/mo estimate; varies by cluster size."),
        ("Defender for App Service", "App Service Instance", app_services, 14.60, app_services * 14.60, None, None,
         "Threat detection for web apps and APIs."),
        ("Defender for Key Vault", "Key Vault", key_vaults, 2.00, key_vaults * 2.00, None, None,
         "Est. ~$0.02/10K ops; $2/vault/mo at moderate use."),
        ("Defender for SQL", "SQL Server / MI", sql, 15.00, sql * 15.00, None, None,
         "Est. $0.015/vCore/hr; $15/server at ~4 vCores avg."),
    ]

    p1_grand_total = 0.0
    p2_grand_total = 0.0
    row_num = 4
    for plan, unit, count, p1_price, p1_total, p2_price, p2_total, notes in plan_rows:
        if count == 0:
            continue
        ws.cell(row=row_num, column=1).value = plan
        ws.cell(row=row_num, column=2).value = unit
        ws.cell(row=row_num, column=3).value = count
        ws.cell(row=row_num, column=4).value = p1_price
        ws.cell(row=row_num, column=5).value = round(p1_total, 2)
        ws.cell(row=row_num, column=6).value = p2_price
        ws.cell(row=row_num, column=7).value = round(p2_total, 2) if p2_total is not None else None
        ws.cell(row=row_num, column=8).value = notes
        p1_grand_total += p1_total
        p2_grand_total += (p2_total if p2_total is not None else p1_total)
        row_num += 1

    # Grand total row
    if row_num > 4:
        from openpyxl.styles import Border, Side
        top_border = Border(top=Side(style="medium"))
        for col in range(1, 9):
            ws.cell(row=row_num, column=col).border = top_border
        ws.cell(row=row_num, column=1).value = "TOTAL MONTHLY ESTIMATE"
        ws.cell(row=row_num, column=1).font = Font(bold=True)
        ws.cell(row=row_num, column=5).value = round(p1_grand_total, 2)
        ws.cell(row=row_num, column=5).font = Font(bold=True)
        ws.cell(row=row_num, column=7).value = round(p2_grand_total, 2)
        ws.cell(row=row_num, column=7).font = Font(bold=True)
        ws.cell(row=row_num, column=8).value = "If all billable resources enrolled. EA/MCA discounts not included."
        ws.cell(row=row_num, column=8).fill = PatternFill("solid", fgColor="DAEEF3")
        for col in [1, 5, 7, 8]:
            ws.cell(row=row_num, column=col).fill = PatternFill("solid", fgColor="DAEEF3")

    _col_widths(ws, [32, 22, 14, 24, 24, 24, 24, 60])


def _write_defender_posture_sheet(wb, scan_data: dict):
    """Defender for Cloud plan posture per subscription (Microsoft.Security/pricings)."""
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    ws = wb.create_sheet("DefenderPosture")

    # Title + reference row
    ws.merge_cells(f"A1:{get_column_letter(7)}1")
    title_cell = ws["A1"]
    title_cell.value = (
        "Defender for Cloud — Plan Posture (Microsoft.Security/pricings). "
        "VMs/VMSS/Arc are analysed per-resource (see DefenderServersCoverage); "
        "other workloads are subscription-level. Requires Reader RBAC."
    )
    title_cell.font = Font(italic=True, size=9, color="1F4E79")
    title_cell.fill = PatternFill("solid", fgColor="D6E4F0")
    title_cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 30

    headers = [
        "Subscription ID", "Defender Plan", "Tier (Standard=On)",
        "Enabled", "Sub-plan", "Resources Coverage", "Enabled Extensions",
    ]
    for col, h in enumerate(headers, 1):
        ws.cell(row=2, column=col).value = h
    _header_style(ws, 2, len(headers))
    _auto_filter(ws, 2, len(headers))
    _freeze(ws, 3)

    sub_names = {
        s.get("subscriptionId", ""): s.get("displayName", "")
        for s in scan_data.get("subscriptions", [])
    }
    _ = sub_names  # subscription name available if a column is added later

    posture = scan_data.get("defender_posture", [])
    posture_sorted = sorted(
        posture,
        key=lambda p: (p.get("subscriptionId", ""), p.get("planDisplayName", "")),
    )
    for row, p in enumerate(posture_sorted, 3):
        ws.cell(row=row, column=1).value = p.get("subscriptionId", "")
        ws.cell(row=row, column=2).value = p.get("planDisplayName", p.get("plan", ""))
        ws.cell(row=row, column=3).value = p.get("pricingTier", "")
        ws.cell(row=row, column=4).value = "Yes" if p.get("enabled") else "No"
        ws.cell(row=row, column=5).value = p.get("subPlan", "")
        ws.cell(row=row, column=6).value = p.get("resourcesCoverageStatus", "")
        ws.cell(row=row, column=7).value = ", ".join(p.get("enabledExtensions", []) or [])
        fill = PatternFill("solid", fgColor=COLOR_OK) if p.get("enabled") else PatternFill("solid", fgColor=COLOR_LOW)
        ws.cell(row=row, column=4).fill = fill

    _col_widths(ws, [38, 38, 18, 10, 18, 22, 45])


def _write_defender_servers_coverage_sheet(wb, scan_data: dict):
    """Per-resource Defender for Servers coverage for VMs / VMSS / Arc Machines."""
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
    from collectors.defender_posture import build_servers_resource_coverage

    rows = build_servers_resource_coverage(
        scan_data.get("defender_posture", []),
        scan_data.get("resources_by_type", {}),
    )
    if not rows:
        return

    ws = wb.create_sheet("DefenderServersCoverage")

    ws.merge_cells(f"A1:{get_column_letter(8)}1")
    title_cell = ws["A1"]
    title_cell.value = (
        "Defender for Servers — Individual Resource Coverage (per-resource). "
        "Derived from the Defender for Servers plan tier of each subscription; "
        "VM-level overrides may apply where coverage is PartiallyCovered."
    )
    title_cell.font = Font(italic=True, size=9, color="1F4E79")
    title_cell.fill = PatternFill("solid", fgColor="D6E4F0")
    title_cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 30

    headers = [
        "Resource", "Type", "Covered", "Plan Tier",
        "Sub-plan", "Location", "Resource Group", "Subscription ID",
    ]
    for col, h in enumerate(headers, 1):
        ws.cell(row=2, column=col).value = h
    _header_style(ws, 2, len(headers))
    _auto_filter(ws, 2, len(headers))
    _freeze(ws, 3)

    rows_sorted = sorted(rows, key=lambda r: (not r.get("covered"), r.get("name", "")))
    for row, r in enumerate(rows_sorted, 3):
        ws.cell(row=row, column=1).value = r.get("name", "")
        ws.cell(row=row, column=2).value = r.get("type", "").split("/")[-1]
        ws.cell(row=row, column=3).value = "Yes" if r.get("covered") else "No"
        ws.cell(row=row, column=4).value = r.get("serversPlanTier", "")
        ws.cell(row=row, column=5).value = r.get("subPlan", "")
        ws.cell(row=row, column=6).value = r.get("location", "")
        ws.cell(row=row, column=7).value = r.get("resourceGroup", "")
        ws.cell(row=row, column=8).value = r.get("subscriptionId", "")
        fill = PatternFill("solid", fgColor=COLOR_OK) if r.get("covered") else PatternFill("solid", fgColor=COLOR_HIGH)
        ws.cell(row=row, column=3).fill = fill

    _col_widths(ws, [40, 22, 10, 14, 16, 16, 28, 38])


def _write_defender_coverage_gap_sheet(wb, scan_data: dict):
    """B2 — Coverage gap & cost to protect (inventory × real plan posture)."""
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
    from collectors.defender_pricing import compute_coverage_gap, total_gap_monthly_cost, pricing_source_label

    gap_rows = compute_coverage_gap(
        scan_data.get("resources_by_type", {}),
        scan_data.get("defender_posture", []),
    )
    if not gap_rows:
        return
    price_src = pricing_source_label(gap_rows)

    ws = wb.create_sheet("DefenderCoverageGap")

    ws.merge_cells(f"A1:{get_column_letter(8)}1")
    title_cell = ws["A1"]
    title_cell.value = "Defender for Cloud — Coverage Gap & Cost to Protect"
    title_cell.font = Font(bold=True, size=13, color="FFFFFF")
    title_cell.fill = PatternFill("solid", fgColor="1F4E79")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 26

    ws.merge_cells(f"A2:{get_column_letter(8)}2")
    disc_cell = ws["A2"]
    disc_cell.value = (
        "Billable units cross-referenced with the real plan posture (Microsoft.Security/pricings). "
        "'Unprotected' = units in subscriptions where the plan is NOT enabled. "
        f"Prices: {price_src}; EA/MCA/CSP discounts, free tiers, and usage-based plans (e.g. Cosmos DB) not reflected."
    )
    disc_cell.font = Font(italic=True, size=9, color="7F7F7F")
    disc_cell.fill = PatternFill("solid", fgColor="FFF2CC")
    disc_cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws.row_dimensions[2].height = 30

    headers = [
        "Defender Plan", "Billable Unit", "Total Units",
        "Protected", "Unprotected (gap)", "Price/Unit/Mo (USD)",
        "Gap Monthly Cost (USD)", "Notes",
    ]
    for col, h in enumerate(headers, 1):
        ws.cell(row=3, column=col).value = h
    _header_style(ws, 3, len(headers))
    _auto_filter(ws, 3, len(headers))
    _freeze(ws, 4)

    row_num = 4
    for r in gap_rows:
        ws.cell(row=row_num, column=1).value = r["plan"]
        ws.cell(row=row_num, column=2).value = r["unit"]
        ws.cell(row=row_num, column=3).value = r["total_units"]
        ws.cell(row=row_num, column=4).value = r["protected_units"]
        ws.cell(row=row_num, column=5).value = r["unprotected_units"]
        ws.cell(row=row_num, column=6).value = r["price"] if r["priceable"] else "Usage-based"
        ws.cell(row=row_num, column=7).value = r["gap_monthly_cost"] if r["priceable"] else None
        ws.cell(row=row_num, column=8).value = r["notes"]
        if r["unprotected_units"] > 0:
            ws.cell(row=row_num, column=5).fill = PatternFill("solid", fgColor=COLOR_HIGH)
        else:
            ws.cell(row=row_num, column=5).fill = PatternFill("solid", fgColor=COLOR_OK)
        row_num += 1

    # Grand total row
    total_units = sum(r["total_units"] for r in gap_rows)
    total_unprotected = sum(r["unprotected_units"] for r in gap_rows)
    total_protected = total_units - total_unprotected
    gap_cost = total_gap_monthly_cost(gap_rows)
    top_border = Border(top=Side(style="medium"))
    for col in range(1, 9):
        ws.cell(row=row_num, column=col).border = top_border
        ws.cell(row=row_num, column=col).fill = PatternFill("solid", fgColor="DAEEF3")
    ws.cell(row=row_num, column=1).value = "TOTAL"
    ws.cell(row=row_num, column=1).font = Font(bold=True)
    ws.cell(row=row_num, column=3).value = total_units
    ws.cell(row=row_num, column=4).value = total_protected
    ws.cell(row=row_num, column=5).value = total_unprotected
    ws.cell(row=row_num, column=7).value = round(gap_cost, 2)
    ws.cell(row=row_num, column=7).font = Font(bold=True)
    ws.cell(row=row_num, column=8).value = f"Monthly cost to protect all unprotected units (prices: {price_src})."

    _col_widths(ws, [34, 22, 12, 12, 18, 20, 24, 55])


