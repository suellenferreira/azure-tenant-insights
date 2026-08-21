import unittest

import openpyxl

from writers.excel_writer import _write_index_sheet, _write_resource_type_sheets


class ExcelResourceTypeContextTests(unittest.TestCase):
    def test_index_includes_resource_type_description_and_count(self) -> None:
        workbook = openpyxl.Workbook()
        workbook.active.title = "Overview"
        workbook.create_sheet("Capacities")
        sheet_map = {"microsoft.powerbidedicated/capacities": "Capacities"}
        scan_data = {
            "resources_by_type": {
                "microsoft.powerbidedicated/capacities": [{"name": "capacity-a"}]
            }
        }

        _write_index_sheet(workbook, scan_data, sheet_map)
        index = workbook["Index"]

        self.assertEqual(
            [index.cell(row=2, column=column).value for column in range(1, 7)],
            ["#", "Sheet", "Resource Type", "Description", "Resource Count", "Go"],
        )
        self.assertEqual(index["C3"].value, "microsoft.powerbidedicated/capacities")
        self.assertIn("Power BI Embedded", index["D3"].value)
        self.assertEqual(index["E3"].value, 1)

    def test_dedicated_resource_sheet_includes_resource_type_after_name(self) -> None:
        workbook = openpyxl.Workbook()
        workbook.remove(workbook.active)
        resource_type = "microsoft.compute/virtualmachines"
        scan_data = {
            "resources_by_type": {
                resource_type: [
                    {
                        "name": "vm-a",
                        "type": resource_type,
                        "resourceGroup": "rg-a",
                        "subscriptionId": "sub-a",
                        "location": "eastus",
                    }
                ]
            },
            "subscriptions": [{"subscriptionId": "sub-a", "displayName": "Subscription A"}],
        }

        _write_resource_type_sheets(
            workbook,
            scan_data,
            include_tags=False,
            sheet_map={resource_type: "Virtual Machines"},
        )
        sheet = workbook["Virtual Machines"]

        self.assertEqual(sheet["A1"].value, "Name")
        self.assertEqual(sheet["B1"].value, "Resource Type")
        self.assertEqual(sheet["B2"].value, resource_type)


if __name__ == "__main__":
    unittest.main()